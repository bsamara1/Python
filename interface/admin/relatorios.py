# interface/admin/relatorios.py
import customtkinter as ctk
from tkinter import messagebox, filedialog
from datetime import datetime
import calendar
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from database.database import conectar

MESES_PT = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]


class RelatoriosPage(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, fg_color="#F4F6FB")
        self.metricas = {}
        self.atualizar_dados()
        self.criar_interface()

    # ------------------------------------------------------------------
    # DADOS
    # ------------------------------------------------------------------
    def obter_metricas(self):
        """Extrai métricas reais da BD (usa sempre a ligação central 'conectar()')"""
        try:
            conn = conectar()
            cursor = conn.cursor()

            cursor.execute("SELECT COUNT(*) FROM estudantes")
            total_estudantes = cursor.fetchone()[0]

            cursor.execute("SELECT COUNT(*) FROM bolsas")
            total_bolsas = cursor.fetchone()[0]

            cursor.execute("SELECT COUNT(*) FROM candidaturas")
            total_candidaturas = cursor.fetchone()[0]

            cursor.execute("SELECT COUNT(*) FROM candidaturas WHERE estado = 'Aprovada'")
            aprovadas = cursor.fetchone()[0]

            cursor.execute("SELECT COUNT(*) FROM candidaturas WHERE estado = 'Pendente'")
            pendentes = cursor.fetchone()[0]

            cursor.execute("SELECT COUNT(*) FROM candidaturas WHERE estado = 'Rejeitada'")
            rejeitadas = cursor.fetchone()[0]

            taxa_aprovacao = (aprovadas / total_candidaturas * 100) if total_candidaturas > 0 else 0

            cursor.execute("SELECT SUM(valor) FROM bolsas WHERE estado = 'Ativo'")
            total_investido = cursor.fetchone()[0] or 0

            # Candidaturas por estado (para o gráfico de rosca)
            cursor.execute("SELECT estado, COUNT(*) FROM candidaturas GROUP BY estado")
            por_estado = cursor.fetchall()

            # Candidaturas por mês (para o gráfico de linha) - últimos 12 meses do ano corrente
            cursor.execute("""
                SELECT strftime('%m', data_candidatura), COUNT(*)
                FROM candidaturas
                WHERE data_candidatura IS NOT NULL
                GROUP BY strftime('%m', data_candidatura)
            """)
            por_mes_raw = dict(cursor.fetchall())

            conn.close()

            return {
                "total_estudantes": total_estudantes,
                "total_bolsas": total_bolsas,
                "total_candidaturas": total_candidaturas,
                "aprovadas": aprovadas,
                "pendentes": pendentes,
                "rejeitadas": rejeitadas,
                "taxa_aprovacao": taxa_aprovacao,
                "total_investido": total_investido,
                "por_estado": por_estado,
                "por_mes_raw": por_mes_raw,
            }
        except Exception as e:
            print(f"Erro ao obter métricas: {e}")
            return {}

    def atualizar_dados(self):
        self.metricas = self.obter_metricas()

    # ------------------------------------------------------------------
    # INTERFACE
    # ------------------------------------------------------------------
    def criar_interface(self):
        # CABEÇALHO
        frame_topo = ctk.CTkFrame(self, fg_color="transparent")
        frame_topo.pack(fill="x", padx=30, pady=(25, 15))

        frame_titulo = ctk.CTkFrame(frame_topo, fg_color="transparent")
        frame_titulo.pack(side="left", fill="x", expand=True)

        ctk.CTkLabel(frame_titulo, text="Relatórios & Estatísticas", font=("Segoe UI", 24, "bold"),
                     text_color="#142850").pack(anchor="w")
        ctk.CTkLabel(frame_titulo, text="Análise de dados, métricas em tempo real e exportação de relatórios.",
                     font=("Segoe UI", 13), text_color="#6B7280").pack(anchor="w", pady=(4, 0))

        frame_botoes = ctk.CTkFrame(frame_topo, fg_color="transparent")
        frame_botoes.pack(side="right")

        ctk.CTkButton(
            frame_botoes, text="📄  Exportar PDF", fg_color="#1D4ED8", hover_color="#1E40AF",
            font=("Segoe UI", 12, "bold"), height=38, width=150, command=self.exportar_pdf
        ).pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            frame_botoes, text="📊  Exportar Excel", fg_color="#10B981", hover_color="#059669",
            font=("Segoe UI", 12, "bold"), height=38, width=160, command=self.exportar_excel
        ).pack(side="left")

        # MÉTRICAS (2 linhas x 4 colunas)
        frame_metricas = ctk.CTkFrame(self, fg_color="transparent")
        frame_metricas.pack(fill="x", padx=20, pady=(5, 15))
        frame_metricas.grid_columnconfigure((0, 1, 2, 3), weight=1)

        m = self.metricas
        cards = [
            ("Total Estudantes", str(m.get("total_estudantes", 0)), "#EEF4FF", "#1E40AF"),
            ("Total Bolsas", str(m.get("total_bolsas", 0)), "#ECFFF0", "#03543F"),
            ("Total Candidaturas", str(m.get("total_candidaturas", 0)), "#FFF8EA", "#92400E"),
            ("Aprovadas", str(m.get("aprovadas", 0)), "#F6EEFF", "#6B21A8"),
            ("Pendentes", str(m.get("pendentes", 0)), "#FFF3E0", "#B45309"),
            ("Rejeitadas", str(m.get("rejeitadas", 0)), "#FEECEC", "#B91C1C"),
            ("Taxa de Aprovação", f"{m.get('taxa_aprovacao', 0):.1f}%", "#E6FBF8", "#0F766E"),
            ("Valor Total (CVE)", f"{m.get('total_investido', 0):,.0f}".replace(",", "."), "#EEF4FF", "#1E3A8A"),
        ]

        for i, (titulo, valor, cor_fundo, cor_texto) in enumerate(cards):
            linha, coluna = divmod(i, 4)
            self.criar_mini_card(frame_metricas, linha, coluna, titulo, valor, cor_fundo, cor_texto)

        # GRÁFICOS
        container_graficos = ctk.CTkFrame(self, fg_color="transparent")
        container_graficos.pack(fill="both", expand=True, padx=20, pady=15)
        container_graficos.grid_columnconfigure((0, 1), weight=1)
        container_graficos.grid_rowconfigure(0, weight=1)

        # Gráfico de linha (candidaturas por mês)
        card_esq = ctk.CTkFrame(container_graficos, fg_color="white", corner_radius=12, border_width=1,
                                 border_color="#E5E7EB")
        card_esq.grid(row=0, column=0, sticky="nsew", padx=(0, 10), pady=5)

        ctk.CTkLabel(card_esq, text="Candidaturas por Mês", font=("Segoe UI", 14, "bold"),
                     text_color="#142850").pack(padx=15, pady=(15, 5), anchor="w")

        self.fig_linha = Figure(figsize=(4, 3), dpi=80, facecolor='white')
        self.criar_grafico_linha()
        self.canvas_linha = FigureCanvasTkAgg(self.fig_linha, master=card_esq)
        self.canvas_linha.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)

        # Gráfico de rosca (candidaturas por estado) + legenda
        card_dir = ctk.CTkFrame(container_graficos, fg_color="white", corner_radius=12, border_width=1,
                                 border_color="#E5E7EB")
        card_dir.grid(row=0, column=1, sticky="nsew", padx=(10, 0), pady=5)

        ctk.CTkLabel(card_dir, text="Candidaturas por Estado", font=("Segoe UI", 14, "bold"),
                     text_color="#142850").pack(padx=15, pady=(15, 5), anchor="w")

        frame_grafico_estado = ctk.CTkFrame(card_dir, fg_color="transparent")
        frame_grafico_estado.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        self.fig_pizza = Figure(figsize=(3, 3), dpi=80, facecolor='white')
        self.criar_grafico_pizza()
        self.canvas_pizza = FigureCanvasTkAgg(self.fig_pizza, master=frame_grafico_estado)
        self.canvas_pizza.get_tk_widget().pack(side="left", fill="both", expand=True)

        self.frame_legenda = ctk.CTkFrame(frame_grafico_estado, fg_color="transparent")
        self.frame_legenda.pack(side="left", fill="y", padx=(10, 0))
        self.criar_legenda_estado()

    def criar_mini_card(self, parent, row, col, titulo, valor, cor_fundo, cor_texto):
        card = ctk.CTkFrame(parent, fg_color=cor_fundo, corner_radius=12, height=85)
        card.grid(row=row, column=col, padx=10, pady=8, sticky="nsew")
        card.pack_propagate(False)

        ctk.CTkLabel(card, text=valor, font=("Segoe UI", 20, "bold"), text_color=cor_texto
                     ).place(relx=0.5, rely=0.35, anchor="center")
        ctk.CTkLabel(card, text=titulo, font=("Segoe UI", 11, "bold"), text_color="#6B7280"
                     ).place(relx=0.5, rely=0.75, anchor="center")

    # ------------------------------------------------------------------
    # GRÁFICOS
    # ------------------------------------------------------------------
    def criar_grafico_linha(self):
        """Gráfico de linha com candidaturas por mês (Jan a Dez do ano corrente)"""
        try:
            por_mes_raw = self.metricas.get("por_mes_raw", {}) or {}
            valores = [por_mes_raw.get(f"{i:02d}", 0) for i in range(1, 13)]

            ax = self.fig_linha.add_subplot(111)
            ax.plot(range(12), valores, marker='o', color='#1A5CFF', linewidth=2, markersize=5)
            ax.fill_between(range(12), valores, alpha=0.25, color='#1A5CFF')
            ax.set_xticks(range(12))
            ax.set_xticklabels(MESES_PT, rotation=45, fontsize=8)
            ax.grid(True, alpha=0.3)
            ax.set_ylim(bottom=0)
            self.fig_linha.tight_layout()

        except Exception as e:
            print(f"Erro ao criar gráfico de linha: {e}")
            ax = self.fig_linha.add_subplot(111)
            ax.text(0.5, 0.5, "Sem dados", ha='center', va='center', transform=ax.transAxes)

    def criar_grafico_pizza(self):
        """Gráfico de rosca (donut) com candidaturas por estado"""
        try:
            dados = self.metricas.get("por_estado", []) or []
            cor_por_estado = {"Aprovada": "#10B981", "Pendente": "#F59E0B", "Rejeitada": "#EF4444"}

            ax = self.fig_pizza.add_subplot(111)

            if dados:
                estados = [d[0] for d in dados]
                valores = [d[1] for d in dados]
                cores = [cor_por_estado.get(e, "#9CA3AF") for e in estados]

                ax.pie(
                    valores, colors=cores, startangle=90,
                    wedgeprops=dict(width=0.42, edgecolor='white')
                )
            else:
                ax.text(0.5, 0.5, "Sem dados", ha='center', va='center', transform=ax.transAxes)
                ax.axis('off')

            self.fig_pizza.tight_layout()

        except Exception as e:
            print(f"Erro ao criar gráfico de pizza: {e}")

    def criar_legenda_estado(self):
        """Legenda manual (cor + estado + percentagem) ao lado da rosca"""
        dados = self.metricas.get("por_estado", []) or []
        cor_por_estado = {"Aprovada": "#10B981", "Pendente": "#F59E0B", "Rejeitada": "#EF4444"}
        total = sum(v for _, v in dados) or 1

        if not dados:
            ctk.CTkLabel(self.frame_legenda, text="Sem dados", font=("Segoe UI", 11),
                         text_color="#9CA3AF").pack(anchor="w", pady=4)
            return

        for estado, valor in dados:
            linha = ctk.CTkFrame(self.frame_legenda, fg_color="transparent")
            linha.pack(anchor="w", pady=4)

            ctk.CTkLabel(linha, text="●", font=("Segoe UI", 14),
                         text_color=cor_por_estado.get(estado, "#9CA3AF")).pack(side="left")

            percentagem = valor / total * 100
            ctk.CTkLabel(linha, text=f"{estado}s ({percentagem:.0f}%)", font=("Segoe UI", 11),
                         text_color="#374151").pack(side="left", padx=(5, 0))

    # ------------------------------------------------------------------
    # EXPORTAÇÃO
    # ------------------------------------------------------------------
    def exportar_pdf(self):
        try:
            caminho = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                initialfile=f"Relatório_Candidaturas_{datetime.now().strftime('%Y%m%d')}.pdf"
            )
            if not caminho:
                return

            doc = SimpleDocTemplate(caminho, pagesize=letter)
            story = [Paragraph("RELATÓRIO DE CANDIDATURAS", getSampleStyleSheet()['Heading1']), Spacer(1, 0.3 * inch)]

            m = self.metricas
            metricas_data = [
                ["Métrica", "Valor"],
                ["Total de Estudantes", str(m.get("total_estudantes", "N/A"))],
                ["Total de Bolsas", str(m.get("total_bolsas", "N/A"))],
                ["Total de Candidaturas", str(m.get("total_candidaturas", "N/A"))],
                ["Candidaturas Aprovadas", str(m.get("aprovadas", "N/A"))],
                ["Candidaturas Pendentes", str(m.get("pendentes", "N/A"))],
                ["Candidaturas Rejeitadas", str(m.get("rejeitadas", "N/A"))],
                ["Taxa de Aprovação", f"{m.get('taxa_aprovacao', 0):.1f}%"],
                ["Total Investido (CVE)", f"{m.get('total_investido', 0):,.0f}".replace(",", ".")],
            ]

            table = Table(metricas_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            doc.build(story)

            messagebox.showinfo("Sucesso", f"Relatório PDF exportado com sucesso!\n{caminho}")

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar PDF: {str(e)}")

    def exportar_excel(self):
        try:
            caminho = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"Relatório_Candidaturas_{datetime.now().strftime('%Y%m%d')}.xlsx"
            )
            if not caminho:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Relatório"

            header_fill = PatternFill(start_color="1A5CFF", end_color="1A5CFF", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

            headers = ["Métrica", "Valor"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal="center")

            m = self.metricas
            dados = [
                ["Total de Estudantes", m.get("total_estudantes", "N/A")],
                ["Total de Bolsas", m.get("total_bolsas", "N/A")],
                ["Total de Candidaturas", m.get("total_candidaturas", "N/A")],
                ["Candidaturas Aprovadas", m.get("aprovadas", "N/A")],
                ["Candidaturas Pendentes", m.get("pendentes", "N/A")],
                ["Candidaturas Rejeitadas", m.get("rejeitadas", "N/A")],
                ["Taxa de Aprovação", f"{m.get('taxa_aprovacao', 0):.1f}%"],
                ["Total Investido (CVE)", m.get("total_investido", 0)],
            ]

            for row_idx, row_data in enumerate(dados, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")

            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20

            wb.save(caminho)
            messagebox.showinfo("Sucesso", f"Relatório Excel exportado com sucesso!\n{caminho}")

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar Excel: {str(e)}")
