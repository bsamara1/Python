import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. Inicializar o Livro de Estilo
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remove a aba padrão

# Criar as abas do sistema SIBES
ws_dash = wb.create_sheet(title="Dashboard")
ws_estudantes = wb.create_sheet(title="Gestão de Estudantes")
ws_bolsas = wb.create_sheet(title="Gestão de Bolsas")
ws_candidaturas = wb.create_sheet(title="Candidaturas")

# 2. Definição da Paleta de Cores (Profissional e Executiva)
PRIMARY_DARK = "1A365D"   
ACCENT_BLUE = "2B6CB0"    
ACCENT_GREEN = "2F855A"   
BORDER_GRAY = "E2E8F0"    

font_title = Font(name="Arial", size=16, bold=True, color="FFFFFF")
font_section = Font(name="Arial", size=13, bold=True, color=PRIMARY_DARK)
font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
font_data = Font(name="Arial", size=11, color="000000")
font_bold_data = Font(name="Arial", size=11, bold=True, color="000000")

fill_menu = PatternFill(start_color=PRIMARY_DARK, end_color=PRIMARY_DARK, fill_type="solid")
fill_header = PatternFill(start_color=ACCENT_BLUE, end_color=ACCENT_BLUE, fill_type="solid")
fill_blue_card = PatternFill(start_color="EBF8FF", end_color="EBF8FF", fill_type="solid")
fill_green_card = PatternFill(start_color="E6FFFA", end_color="E6FFFA", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color=BORDER_GRAY),
    right=Side(style="thin", color=BORDER_GRAY),
    top=Side(style="thin", color=BORDER_GRAY),
    bottom=Side(style="thin", color=BORDER_GRAY)
)

align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

# ----------------------------------------------------
# ABA 1: DASHBOARD DE MÉTRICAS
# ----------------------------------------------------
ws_dash.views.sheetView[0].showGridLines = True

ws_dash.merge_cells("A1:G2")
ws_dash["A1"] = "SIBES - Sistema Inteligente de Bolsas de Estudo Sustentáveis"
ws_dash["A1"].font = font_title
ws_dash["A1"].fill = fill_menu
ws_dash["A1"].alignment = align_center

# Indicadores Rápidos
ws_dash.merge_cells("A4:B5")
ws_dash["A4"] = "120\nEstudantes Registados"
ws_dash["A4"].font = Font(name="Arial", size=12, bold=True, color="2B6CB0")
ws_dash["A4"].fill = fill_blue_card
ws_dash["A4"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws_dash.merge_cells("C4:D5")
ws_dash["C4"] = "15\nBolsas Disponíveis"
ws_dash["C4"].font = Font(name="Arial", size=12, bold=True, color=ACCENT_GREEN)
ws_dash["C4"].fill = fill_green_card
ws_dash["C4"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws_dash.merge_cells("E4:F5")
ws_dash["E4"] = "85\nCandidaturas Submetidas"
ws_dash["E4"].font = Font(name="Arial", size=12, bold=True, color="B7791F")
ws_dash["E4"].fill = PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid")
ws_dash["E4"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Desenhar bordas nos cards
for row in ws_dash["A4:F5"]:
    for cell in row:
        cell.border = thin_border

ws_dash["A7"] = "Visão Geral de Candidaturas Pendentes"
ws_dash["A7"].font = font_section

headers_dash = ["ID", "Estudante", "Bolsa Pretendida", "Data", "Estado", "Ação da Secretaria"]
for col_num, header in enumerate(headers_dash, 1):
    cell = ws_dash.cell(row=8, column=col_num)
    cell.value = header
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = thin_border

dash_data = [
    [1, "João Silva", "Bolsa de Mérito", "2026-07-01", "Pendente", "Validar Documentação"],
    [2, "Ana Santos", "Bolsa Social", "2026-07-02", "Aprovada", "Finalizado"],
    [3, "Carlos Lima", "Bolsa Desporto", "2026-07-03", "Pendente", "Aguardar Parecer Técnico"]
]

for r_idx, row_data in enumerate(dash_data, 9):
    for c_idx, val in enumerate(row_data, 1):
        cell = ws_dash.cell(row=r_idx, column=c_idx)
        cell.value = val
        cell.font = font_data
        cell.border = thin_border
        cell.alignment = align_center if c_idx in [1, 4, 5] else align_left

# ----------------------------------------------------
# ABA 2: GESTÃO DE ESTUDANTES
# ----------------------------------------------------
ws_estudantes.views.sheetView[0].showGridLines = True
ws_estudantes["A1"] = "Base de Dados de Estudantes Registados"
ws_estudantes["A1"].font = font_section

headers_estudantes = ["ID", "Nome do Estudante", "Email", "Curso", "Média", "Rendimento Familiar (CVE)"]
for col_num, header in enumerate(headers_estudantes, 1):
    cell = ws_estudantes.cell(row=3, column=col_num)
    cell.value = header
    cell.font = font_header
    cell.fill = fill_menu
    cell.alignment = align_center
    cell.border = thin_border

estudantes_data = [
    [1, "João Silva", "joao@email.com", "Eng. Informática", 15.5, 35000],
    [2, "Ana Santos", "ana@email.com", "Gestão", 16.0, 28000],
    [3, "Carlos Lima", "carlos@email.com", "Contabilidade", 14.2, 40000]
]

for r_idx, row_data in enumerate(estudantes_data, 4):
    for c_idx, val in enumerate(row_data, 1):
        cell = ws_estudantes.cell(row=r_idx, column=c_idx)
        cell.value = val
        cell.font = font_data
        cell.border = thin_border
        if c_idx == 1: cell.alignment = align_center
        elif c_idx in [2,3,4]: cell.alignment = align_left
        elif c_idx == 5: 
            cell.alignment = align_right
            cell.number_format = "0.0"
        elif c_idx == 6:
            cell.alignment = align_right
            cell.number_format = "#,##0"

# ----------------------------------------------------
# ABA 3: GESTÃO DE BOLSAS
# ----------------------------------------------------
ws_bolsas.views.sheetView[0].showGridLines = True
ws_bolsas["A1"] = "Critérios e Parâmetros Globais de Bolsas"
ws_bolsas["A1"].font = font_section

headers_bolsas = ["ID", "Nome da Bolsa", "Valor (CVE)", "Média Mínima", "Rendimento Máximo (CVE)"]
for col_num, header in enumerate(headers_bolsas, 1):
    cell = ws_bolsas.cell(row=3, column=col_num)
    cell.value = header
    cell.font = font_header
    cell.fill = fill_menu
    cell.alignment = align_center
    cell.border = thin_border

bolsas_data = [
    [1, "Bolsa de Mérito", 50000, 15.0, 50000],
    [2, "Bolsa Social", 40000, 12.0, 35000]
]

for r_idx, row_data in enumerate(bolsas_data, 4):
    for c_idx, val in enumerate(row_data, 1):
        cell = ws_bolsas.cell(row=r_idx, column=c_idx)
        cell.value = val
        cell.font = font_data
        cell.border = thin_border
        if c_idx == 1: cell.alignment = align_center
        elif c_idx == 2: cell.alignment = align_left
        elif c_idx == 4:
            cell.alignment = align_right
            cell.number_format = "0.0"
        elif c_idx in [3, 5]:
            cell.alignment = align_right
            cell.number_format = "#,##0"

# ----------------------------------------------------
# ABA 4: CANDIDATURAS COM VALIDAÇÃO AUTOMÁTICA
# ----------------------------------------------------
ws_candidaturas.views.sheetView[0].showGridLines = True
ws_candidaturas["A1"] = "Cruzamento Inteligente de Dados"
ws_candidaturas["A1"].font = font_section

headers_candidaturas = [
    "ID", "Estudante", "Bolsa", "Média Aluno", "Média Mínima", 
    "Rendimento Aluno", "Rendimento Máx", "Elegibilidade"
]

for col_num, header in enumerate(headers_candidaturas, 1):
    cell = ws_candidaturas.cell(row=3, column=col_num)
    cell.value = header
    cell.font = font_header
    cell.fill = fill_menu
    cell.alignment = align_center
    cell.border = thin_border

cand_base = [
    [1, "João Silva", "Bolsa de Mérito", "='Gestão de Estudantes'!E4", "='Gestão de Bolsas'!D4", "='Gestão de Estudantes'!F4", "='Gestão de Bolsas'!E4"],
    [2, "Ana Santos", "Bolsa Social", "='Gestão de Estudantes'!E5", "='Gestão de Bolsas'!D5", "='Gestão de Estudantes'!F5", "='Gestão de Bolsas'!E5"]
]

for r_idx, row_data in enumerate(cand_base, 4):
    for c_idx, val in enumerate(row_data, 1):
        cell = ws_candidaturas.cell(row=r_idx, column=c_idx)
        cell.value = val
        cell.font = font_data
        cell.border = thin_border
        if c_idx == 1: cell.alignment = align_center
        elif c_idx in [2, 3]: cell.alignment = align_left
        elif c_idx in [4, 5]: cell.number_format = "0.0"
        elif c_idx in [6, 7]: cell.number_format = "#,##0"

    # Inserção da fórmula dinâmica
    formula_cell = ws_candidaturas.cell(row=r_idx, column=8)
    formula_cell.value = f'=IF(AND(D{r_idx}>=E{r_idx},F{r_idx}<=G{r_idx}),"Elegível","Não Elegível")'
    formula_cell.font = font_bold_data
    formula_cell.alignment = align_center
    formula_cell.border = thin_border

# Auto-ajuste de colunas
for ws in [ws_dash, ws_estudantes, ws_bolsas, ws_candidaturas]:
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

wb.save("SIBES_Dashboard_Secretaria.xlsx")